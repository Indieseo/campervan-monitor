"""
Data Export Engine
Export competitive intelligence data to Excel, PDF, and CSV formats
"""

import pandas as pd
import sqlite3
import sys
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, Any, Optional
from loguru import logger
import json

# Fix Windows console encoding for emoji support
if sys.platform == 'win32':
    import codecs
    if hasattr(sys.stdout, 'buffer'):
        sys.stdout = codecs.getwriter('utf-8')(sys.stdout.buffer, 'strict')
        sys.stderr = codecs.getwriter('utf-8')(sys.stderr.buffer, 'strict')

# Excel formatting
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.chart import LineChart, BarChart, Reference
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    logger.warning("openpyxl not installed. Excel export disabled.")

# PDF generation
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False
    logger.warning("reportlab not installed. PDF export disabled.")

BASE_DIR = Path(__file__).parent.resolve()
EXPORT_DIR = BASE_DIR / "exports"
EXPORT_DIR.mkdir(exist_ok=True)


class DataExporter:
    """Export intelligence data in multiple formats"""
    
    def __init__(self, db_path: str = None):
        self.db_path = db_path or str(BASE_DIR / "database" / "campervan_prices.db")
        self.export_dir = EXPORT_DIR
    
    def export_to_excel(self, days: int = 30) -> Optional[str]:
        """
        Export data to formatted Excel with multiple sheets and charts
        """
        if not EXCEL_AVAILABLE:
            logger.error("Excel export not available. Install: pip install openpyxl")
            return None
        
        try:
            conn = sqlite3.connect(self.db_path)
            
            # Get data
            cutoff_date = (datetime.now() - timedelta(days=days)).strftime('%Y-%m-%d')
            
            # Main prices data
            prices_df = pd.read_sql(f"""
                SELECT 
                    company_name,
                    base_price,
                    discount_percentage,
                    scrape_date,
                    vehicle_type,
                    location
                FROM prices
                WHERE scrape_date >= '{cutoff_date}'
                ORDER BY scrape_date DESC
            """, conn)
            
            # Summary by company
            summary_df = pd.read_sql(f"""
                SELECT 
                    company_name,
                    COUNT(*) as data_points,
                    AVG(base_price) as avg_price,
                    MIN(base_price) as min_price,
                    MAX(base_price) as max_price,
                    AVG(CASE WHEN discount_percentage > 0 THEN discount_percentage END) as avg_discount
                FROM prices
                WHERE scrape_date >= '{cutoff_date}'
                GROUP BY company_name
                ORDER BY avg_price ASC
            """, conn)
            
            # Create Excel file
            filename = f"campervan_intel_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
            filepath = self.export_dir / filename
            
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                # Sheet 1: Summary
                summary_df.to_excel(writer, sheet_name='Summary', index=False)
                
                # Sheet 2: Raw Data
                prices_df.to_excel(writer, sheet_name='Raw Data', index=False)
                
                # Sheet 3: Price Trends
                trends_df = prices_df.pivot_table(
                    values='base_price',
                    index='scrape_date',
                    columns='company_name',
                    aggfunc='mean'
                ).reset_index()
                trends_df.to_excel(writer, sheet_name='Price Trends', index=False)
                
                # Format the workbook
                workbook = writer.book
                self._format_excel_sheets(workbook)
            
            conn.close()
            logger.info(f"✅ Excel exported: {filepath}")
            return str(filepath)
            
        except Exception as e:
            logger.error(f"❌ Excel export failed: {e}")
            return None
    
    def _format_excel_sheets(self, workbook):
        """Apply professional formatting to Excel sheets"""
        header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            # Format headers
            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
                cell.border = border
            
            # Auto-adjust column width
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[column_letter].width = adjusted_width
    
    def export_to_pdf(self, days: int = 7) -> Optional[str]:
        """Generate PDF executive report"""
        if not PDF_AVAILABLE:
            logger.error("PDF export not available. Install: pip install reportlab")
            return None
        
        try:
            conn = sqlite3.connect(self.db_path)
            cutoff_date = (datetime.now() - timedelta(days=days)).strftime('%Y-%m-%d')
            
            # Get summary data
            summary_query = f"""
                SELECT 
                    company_name,
                    AVG(base_price) as avg_price,
                    MIN(base_price) as min_price,
                    MAX(base_price) as max_price
                FROM prices
                WHERE scrape_date >= '{cutoff_date}'
                GROUP BY company_name
                ORDER BY avg_price ASC
            """
            summary_data = pd.read_sql(summary_query, conn)
            
            # Create PDF
            filename = f"executive_report_{datetime.now().strftime('%Y%m%d')}.pdf"
            filepath = self.export_dir / filename
            
            doc = SimpleDocTemplate(str(filepath), pagesize=letter)
            story = []
            styles = getSampleStyleSheet()
            
            # Title
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                textColor=colors.HexColor('#2C3E50'),
                spaceAfter=30
            )
            story.append(Paragraph("🚐 Campervan Competitive Intelligence", title_style))
            story.append(Paragraph(f"Executive Report - {datetime.now().strftime('%B %d, %Y')}", styles['Normal']))
            story.append(Spacer(1, 0.3*inch))
            
            # Market Summary
            story.append(Paragraph("📊 Market Summary", styles['Heading2']))
            market_avg = summary_data['avg_price'].mean()
            story.append(Paragraph(f"Average Market Price: €{market_avg:.2f}/night", styles['Normal']))
            story.append(Paragraph(f"Price Range: €{summary_data['min_price'].min():.2f} - €{summary_data['max_price'].max():.2f}", styles['Normal']))
            story.append(Spacer(1, 0.3*inch))
            
            # Competitor Table
            story.append(Paragraph("🏆 Competitor Pricing", styles['Heading2']))
            table_data = [['Company', 'Avg Price', 'Min Price', 'Max Price']]
            for _, row in summary_data.iterrows():
                table_data.append([
                    row['company_name'],
                    f"€{row['avg_price']:.2f}",
                    f"€{row['min_price']:.2f}",
                    f"€{row['max_price']:.2f}"
                ])
            
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2C3E50')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(table)
            
            # Build PDF
            doc.build(story)
            conn.close()
            
            logger.info(f"✅ PDF exported: {filepath}")
            return str(filepath)
            
        except Exception as e:
            logger.error(f"❌ PDF export failed: {e}")
            return None
    
    def export_to_csv(self, days: int = 30) -> Optional[str]:
        """Export raw data to CSV"""
        try:
            conn = sqlite3.connect(self.db_path)
            cutoff_date = (datetime.now() - timedelta(days=days)).strftime('%Y-%m-%d')
            
            df = pd.read_sql(f"""
                SELECT * FROM prices
                WHERE scrape_date >= '{cutoff_date}'
                ORDER BY scrape_date DESC, company_name
            """, conn)
            
            filename = f"campervan_data_{datetime.now().strftime('%Y%m%d')}.csv"
            filepath = self.export_dir / filename
            
            df.to_csv(filepath, index=False)
            conn.close()
            
            logger.info(f"✅ CSV exported: {filepath}")
            return str(filepath)
            
        except Exception as e:
            logger.error(f"❌ CSV export failed: {e}")
            return None
    
    def export_summary_json(self) -> Optional[str]:
        """Export quick summary as JSON"""
        try:
            conn = sqlite3.connect(self.db_path)
            
            summary = {
                'generated_at': datetime.now().isoformat(),
                'market_summary': {},
                'top_competitors': [],
                'recent_alerts': []
            }
            
            # Market summary
            market_data = pd.read_sql("""
                SELECT 
                    AVG(base_price) as avg_price,
                    MIN(base_price) as min_price,
                    MAX(base_price) as max_price,
                    COUNT(DISTINCT company_name) as companies
                FROM prices
                WHERE scrape_date >= date('now', '-7 days')
            """, conn)
            
            summary['market_summary'] = {
                'avg_price': float(market_data['avg_price'].iloc[0]),
                'min_price': float(market_data['min_price'].iloc[0]),
                'max_price': float(market_data['max_price'].iloc[0]),
                'companies_tracked': int(market_data['companies'].iloc[0])
            }
            
            # Top competitors
            top_df = pd.read_sql("""
                SELECT 
                    company_name,
                    AVG(base_price) as avg_price
                FROM prices
                WHERE scrape_date >= date('now', '-7 days')
                GROUP BY company_name
                ORDER BY avg_price ASC
                LIMIT 5
            """, conn)
            
            summary['top_competitors'] = [
                {'company': row['company_name'], 'avg_price': float(row['avg_price'])}
                for _, row in top_df.iterrows()
            ]
            
            filename = f"summary_{datetime.now().strftime('%Y%m%d_%H%M')}.json"
            filepath = self.export_dir / filename
            
            with open(filepath, 'w') as f:
                json.dump(summary, f, indent=2)
            
            conn.close()
            logger.info(f"✅ JSON summary exported: {filepath}")
            return str(filepath)
            
        except Exception as e:
            logger.error(f"❌ JSON export failed: {e}")
            return None
    
    def scheduled_export(self, format: str = 'excel') -> bool:
        """Run scheduled export (called by automation)"""
        logger.info(f"📊 Running scheduled {format} export...")
        
        if format == 'excel':
            result = self.export_to_excel(days=30)
        elif format == 'pdf':
            result = self.export_to_pdf(days=7)
        elif format == 'csv':
            result = self.export_to_csv(days=30)
        else:
            logger.error(f"Unknown format: {format}")
            return False
        
        return result is not None


# Test/CLI usage
if __name__ == "__main__":
    import sys
    
    exporter = DataExporter()
    
    if len(sys.argv) > 1:
        format_type = sys.argv[1].lower()
        
        if format_type == 'excel':
            print("📊 Exporting to Excel...")
            file = exporter.export_to_excel()
        elif format_type == 'pdf':
            print("📄 Exporting to PDF...")
            file = exporter.export_to_pdf()
        elif format_type == 'csv':
            print("📋 Exporting to CSV...")
            file = exporter.export_to_csv()
        elif format_type == 'json':
            print("📦 Exporting to JSON...")
            file = exporter.export_summary_json()
        elif format_type == 'all':
            print("📊 Exporting all formats...")
            exporter.export_to_excel()
            exporter.export_to_pdf()
            exporter.export_to_csv()
            exporter.export_summary_json()
            file = "All formats exported"
        else:
            print("❌ Unknown format. Use: excel, pdf, csv, json, or all")
            sys.exit(1)
        
        if file:
            print(f"✅ Success: {file}")
    else:
        print("Usage: python export_engine.py [excel|pdf|csv|json|all]")
        print("Example: python export_engine.py excel")
